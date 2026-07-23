# -*- coding: utf-8 -*-
"""
Сборка data.json из двух Excel-файлов конкурентного анализа (схема v2, 35 колонок).
Приточные (supply) и приточно-вытяжные (pvu) — единый массив с тегом type.
Методика баллов — LOGIC.md. Файлы ищутся рядом с webapp/ и на уровень выше.

Запуск:  python tools/build_data.py
"""
import openpyxl, json, re, os, statistics

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILES = [
    ("pvu",    "Конкурентный_анализ_ПВУ_Shuft_и_конкуренты.xlsx"),
    ("supply", "Конкурентный_анализ_ПУ_Shuft_и_конкуренты.xlsx"),
]

# A..AI (35 колонок) -> ключ модели (None = пропустить)
COLS = [
    "brand",        # 0  Бренд
    "camp",         # 1  Наши / Конкурент
    "series",       # 2  Серия
    "name",         # 3  Наименование
    "price",        # 4  Цена, руб
    "url",          # 5  Ссылка на карточку
    "passport",     # 6  Паспорт (источник)
    "flow",         # 7  Расход воздуха, м³/ч
    "pressure",     # 8  Свободный напор, Па
    "dims",         # 9  Размеры Ш×В×Г, мм
    "weight",       # 10 Масса, кг
    "power",        # 11 Питание, ф/В/Гц
    "fan_kw",       # 12 Мощность вентиляторов, кВт
    "heat_kw",      # 13 Мощность нагревателя, кВт
    "ip",           # 14 Степень защиты IP
    "noise",        # 15 Уровень шума, дБ(А)
    "valve",        # 16 Воздушный клапан
    "valve_drive",  # 17 Привод клапана
    "filter_class", # 18 Класс фильтрации
    "recup",        # 19 Рекуперация
    "recup_type",   # 20 Тип рекуператора
    "recup_eff",    # 21 КПД рекуперации, %
    "heater_type",  # 22 Нагреватель
    "heater_elem",  # 23 Нагревательный элемент
    "cooler",       # 24 Охладитель
    "fan_type",     # 25 Тип вентилятора
    "motor",        # 26 Двигатель (AC/EC)
    "auto",         # 27 Автоматика
    "controller",   # 28 Контроллер
    "remote",       # 29 Пульт
    "wifi",         # 30 Wi-Fi
    "vav",          # 31 VAV
    "humidity",     # 32 Управление влажностью
    "co2",          # 33 Датчик CO2
    "extra",        # 34 Примечание
]

LADDER = [("H14",18),("H13",17),("H11",15),("HEPA",15),("E12",14),("E11",13),("EPA",12),
          ("E10",12),("EU9",11),("F9",11),("F8",10),("EU7",9),("F7",9),("F6",8),("EU5",7),
          ("F5",7),("M6",6),("M5",6),("EU4",4),("G4",4),("EU3",3),("G3",3),("G2",2),
          ("ПЫЛЕВОЙ",2),("G1",1)]

PVU_SEG = [(250,"S1"),(400,"S2"),(550,"S3"),(900,"S4"),(1300,"S5"),(1700,"S6"),(2500,"S7"),(float("inf"),"S8")]
SUP_SEG = [(500,"S1"),(800,"S2"),(1400,"S3"),(2500,"S4"),(3500,"S5"),(5000,"S6"),(float("inf"),"S7")]

def s(v):  # safe trimmed string
    return "" if v is None else str(v).strip()

def ints(x):
    return [int(n) for n in re.findall(r"\d+", s(x))]

def price_num(b):
    d = re.sub(r"[^\d]", "", s(b))
    return int(d) if d else None

def flow_max(e):
    xs = ints(e)
    return max(xs) if xs else None

def thickness(g):
    if "запрос" in s(g).lower():
        return None
    xs = [n for n in ints(g) if n >= 50]
    return min(xs) if xs else None

def noise_max(l):
    xs = [n for n in ints(l) if 10 <= n <= 90]
    return max(xs) if xs else None

def eff_pct(u):
    """КПД: в новых файлах число без «%» (78, 54); допускаем и «78%»/«0,78»."""
    sv = s(u).lower()
    if not sv or sv in ("н/д", "нет", "-", "—"):
        return None
    m = re.search(r"\d+(?:[.,]\d+)?", sv)
    if not m:
        return None
    v = float(m.group(0).replace(",", "."))
    if v <= 1.0:      # доля 0.78 -> 78
        v *= 100
    return int(round(v)) if 1 <= v <= 100 else None

def filter_rank(q):
    """Ранг ШТАТНОГО фильтра: опции не учитываем, кириллическую М нормализуем."""
    up = s(q).upper().replace(" ", " ")
    up = up.replace("М", "M")                      # кириллица -> латиница (М5/М6)
    up = re.sub(r"\b([A-Z]{1,4}\d{0,2})\s*[-–—]?\s*ОПЦИ\w*", "", up)   # «F7 опция»
    up = re.split(r"ОПЦ", up)[0]                   # отрезать «(опции: …)» и дальше
    up = up.replace(" ", "")
    best = None
    for key, rank in LADDER:
        if key in up:
            best = rank if best is None else max(best, rank)
    return best

def is_yes(v):
    return s(v).lower().startswith("да")           # «да», «да (web)», «да, сенсорный…»

def func_count(row):
    return sum(1 for c in (row.get("wifi"), row.get("vav"), row.get("humidity"), row.get("co2")) if is_yes(c))

def norm_brand(b):
    bl = s(b).lower()
    if bl in ("rusklimat", "shuft"):
        return "SHUFT"
    return s(b)

def has_filter(filter_class):
    fl = s(filter_class).lower()
    if not fl or fl in ("н/д", "-", "—"):
        return "н/д"
    return "нет" if fl.startswith("нет") else "да"

def segment(flow, table):
    if flow is None:
        return None
    for hi, name in table:
        if flow <= hi:
            return name
    return table[-1][1]

def find_file(fn):
    for base in (ROOT, os.path.dirname(ROOT)):
        p = os.path.join(base, fn)
        if os.path.exists(p):
            return p
    raise FileNotFoundError(fn)

def load(path, typ):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Данные"]
    out = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in r]
        a = s(vals[0]) if len(vals) > 0 else ""
        c = s(vals[3]) if len(vals) > 3 else ""
        if not a and not c:
            continue
        m = {}
        for i, key in enumerate(COLS):
            if key is None or i >= len(vals):
                continue
            m[key] = s(vals[i])
        m["brand"] = norm_brand(m.get("brand"))
        m["type"] = typ
        # parsed numeric inputs
        m["price_num"]   = price_num(m.get("price"))
        m["flow_max"]    = flow_max(m.get("flow"))
        m["thickness"]   = thickness(m.get("dims"))
        m["noise_max"]   = noise_max(m.get("noise"))
        m["filter_rank"] = filter_rank(m.get("filter_class"))
        m["filter"]      = has_filter(m.get("filter_class"))
        m["func_count"]  = func_count(m)
        m["eff"]         = eff_pct(m.get("recup_eff")) if typ == "pvu" else None
        m["segment"]     = segment(m["flow_max"], PVU_SEG if typ == "pvu" else SUP_SEG)
        out.append(m)
    wb.close()
    return out

# ───────────────────────── verification (mirror of rating.js) ─────────────────
SEGMENTAL = ["price_num", "thickness", "noise_max"]   # less is better, per-segment
def score_type(models, typ):
    seg_bounds = {}
    for k in SEGMENTAL:
        for m in models:
            seg = m["segment"]
            v = m.get(k)
            if v is None or seg is None: continue
            seg_bounds.setdefault((k,seg), [v,v])
            b = seg_bounds[(k,seg)]; b[0]=min(b[0],v); b[1]=max(b[1],v)
    glob = {}
    for k in (["filter_rank"] + (["eff"] if typ=="pvu" else [])):
        vs = [m[k] for m in models if m.get(k) is not None]
        glob[k] = (min(vs), max(vs)) if vs else None
    def sc_seg(k,m):
        v=m.get(k)
        if v is None or m["segment"] is None: return 0.0
        b=seg_bounds.get((k,m["segment"]))
        if not b: return 0.0
        lo,hi=b
        if hi==lo: return 50.0
        return max(0.0,min(100.0,100.0*(hi-v)/(hi-lo)))
    def sc_abs(k,m):
        v=m.get(k)
        if v is None: return 0.0
        g=glob.get(k)
        if not g: return 0.0
        lo,hi=g
        if hi==lo: return 50.0
        return max(0.0,min(100.0,100.0*(v-lo)/(hi-lo)))
    for m in models:
        parts={}
        parts["price"]=sc_seg("price_num",m)
        parts["thick"]=sc_seg("thickness",m)
        parts["noise"]=sc_seg("noise_max",m)
        parts["filter"]=sc_abs("filter_rank",m)
        parts["func"]=m["func_count"]/4*100.0
        if typ=="pvu": parts["eff"]=sc_abs("eff",m)
        m["_parts"]=parts
        m["_total"]=round(sum(parts.values())/len(parts))
    return models

def report(models, typ, log):
    log.append(f"\n===== {typ.upper()}  ({len(models)} моделей) =====")
    from collections import Counter
    bc=Counter(m["brand"] for m in models)
    log.append("Бренды: " + ", ".join(f"{k}:{bc[k]}" for k in sorted(bc)))
    segc=Counter(m["segment"] for m in models)
    log.append("Сегменты: " + ", ".join(f"{k}:{segc[k]}" for k in sorted(segc, key=lambda x:(x is None, str(x)))))
    totals=[m["_total"] for m in models]
    log.append(f"ИТОГ: min={min(totals)} median={int(statistics.median(totals))} max={max(totals)} mean={round(statistics.mean(totals),1)}")
    for k,lab in [("price_num","Цена"),("thickness","Толщина"),("noise_max","Шум"),("filter_rank","Фильтр"),("eff","КПД")]:
        if k=="eff" and typ!="pvu": continue
        miss=sum(1 for m in models if m.get(k) is None)
        log.append(f"  н/д {lab}: {miss} ({round(100*miss/len(models))}%)")
    log.append("ТОП-7 по ИТОГ:")
    for m in sorted(models,key=lambda x:-x["_total"])[:7]:
        log.append(f"   {m['_total']:>3}  [{m['segment']}] {m['brand']} · {m['name'][:52]}")
    log.append("Худшие-5 по ИТОГ:")
    for m in sorted(models,key=lambda x:x["_total"])[:5]:
        log.append(f"   {m['_total']:>3}  [{m['segment']}] {m['brand']} · {m['name'][:52]}")

def main():
    all_models=[]
    log=[]
    for typ,fn in FILES:
        ms=load(find_file(fn),typ)
        score_type(ms,typ)
        report(ms,typ,log)
        all_models.extend(ms)
    out=[]
    for i,m in enumerate(all_models):
        m=dict(m); m.pop("_parts",None); m.pop("_total",None)
        m["id"]=i
        out.append(m)
    with open(os.path.join(ROOT,"data.json"),"w",encoding="utf-8") as f:
        json.dump(out,f,ensure_ascii=False,separators=(",",":"))
    log.append(f"\nИТОГО записано в data.json: {len(out)} моделей")
    print("\n".join(log))

if __name__=="__main__":
    main()
